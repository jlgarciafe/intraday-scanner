"""
ta_runner.py
────────────────────────────────────────────────────────────
Automated TA Entry Framework — runs after intraday scanner.
For each qualifying ticker, computes:
  - Trend classification
  - Key support / resistance levels
  - RSI, MACD, ADX
  - Entry zone, Stop Loss, T1, T2, Recommended Exit
  - R/R ratios
  - Verdict: ENTER NOW / WAIT FOR DIP / WAIT FOR BREAKOUT / PASS

Outputs:
  - orders.json   -> machine-readable, broker-ready (top 10 ranked)
  - ta_report.md  -> full audit trail (all tickers)
  - Telegram      -> top 10 only, sorted highest return probability first

Usage:
  python ta_runner.py --tickers AAPL MSFT ARHS
  python ta_runner.py --from-file scan_results.json
"""

import os
import re
import json
import argparse
import math
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
import yfinance as yf

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────

LOOKBACK_DAYS       = 252
MIN_RR_T1           = 1.2
MIN_RR_T2           = 2.0
MAX_STOP_PCT        = 0.25
MIN_STOP_PCT        = 0.05
ADX_TREND_THRESHOLD = 25
TOP_N               = 10     # Top N ranked setups sent to Telegram

TELEGRAM_BOT_TOKEN  = os.getenv("TELEGRAM_BOT_TOKEN", "")
TELEGRAM_CHAT_ID    = os.getenv("TELEGRAM_CHAT_ID", "")
DRY_RUN             = os.getenv("DRY_RUN", "false").lower() == "true"

# ── Position sizing ───────────────────────────────────────────────────────────
# Capital and risk % are read from env so they can be adjusted without code changes.
# Default: €250k capital, 2.5% risk per trade (middle of the 2-3% range).
CAPITAL_EUR = float(os.getenv("CAPITAL_EUR", "250000"))
RISK_PCT    = float(os.getenv("RISK_PCT",    "0.025"))

# ── FX rate cache (populated lazily, one fetch per unique currency per run) ──
_FX_CACHE: dict = {}

def get_eur_fx_rate(currency: str) -> float:
    """
    Returns units of `currency` per 1 EUR.
    e.g. EURKRW=X ≈ 1550  →  1 EUR = 1,550 KRW  →  returns 1550.0
         EURUSD=X ≈ 1.08  →  1 EUR = 1.08 USD   →  returns 1.08
         EUR itself        →  returns 1.0
    Falls back to 1.0 on any network error so sizing still runs.
    """
    if currency == "EUR":
        return 1.0
    if currency in _FX_CACHE:
        return _FX_CACHE[currency]
    try:
        df = yf.download(f"EUR{currency}=X", period="2d", progress=False)
        if isinstance(df.columns, __import__("pandas").MultiIndex):
            df.columns = df.columns.get_level_values(0)
        if not df.empty and float(df["Close"].iloc[-1]) > 0:
            rate = float(df["Close"].iloc[-1])
            _FX_CACHE[currency] = rate
            return rate
    except Exception:
        pass
    _FX_CACHE[currency] = 1.0
    return 1.0


# ─────────────────────────────────────────────────────────────────────────────
# TECHNICAL INDICATORS
# ─────────────────────────────────────────────────────────────────────────────

def compute_rsi(close, period=14):
    delta    = close.diff()
    gain     = delta.clip(lower=0)
    loss     = -delta.clip(upper=0)
    avg_gain = gain.ewm(com=period - 1, min_periods=period).mean()
    avg_loss = loss.ewm(com=period - 1, min_periods=period).mean()
    rs       = avg_gain / avg_loss.replace(0, np.nan)
    rsi      = 100 - (100 / (1 + rs))
    return round(float(rsi.iloc[-1]), 2)


def compute_macd(close):
    ema12  = close.ewm(span=12, adjust=False).mean()
    ema26  = close.ewm(span=26, adjust=False).mean()
    macd   = ema12 - ema26
    signal = macd.ewm(span=9, adjust=False).mean()
    hist   = macd - signal
    return (
        round(float(macd.iloc[-1]), 4),
        round(float(signal.iloc[-1]), 4),
        round(float(hist.iloc[-1]), 4),
    )


def compute_adx(high, low, close, period=14):
    tr = pd.concat([
        high - low,
        (high - close.shift()).abs(),
        (low  - close.shift()).abs()
    ], axis=1).max(axis=1)
    dm_pos = high.diff().clip(lower=0)
    dm_neg = (-low.diff()).clip(lower=0)
    dm_pos = dm_pos.where(dm_pos > dm_neg, 0)
    dm_neg = dm_neg.where(dm_neg > dm_pos, 0)
    tr_s = tr.ewm(com=period - 1, min_periods=period).mean()
    dip  = 100 * dm_pos.ewm(com=period - 1, min_periods=period).mean() / tr_s
    din  = 100 * dm_neg.ewm(com=period - 1, min_periods=period).mean() / tr_s
    dx   = 100 * (dip - din).abs() / (dip + din).replace(0, np.nan)
    adx  = dx.ewm(com=period - 1, min_periods=period).mean()
    return round(float(adx.iloc[-1]), 2)


def compute_atr(high, low, close, period=14):
    tr = pd.concat([
        high - low,
        (high - close.shift()).abs(),
        (low  - close.shift()).abs()
    ], axis=1).max(axis=1)
    return round(float(tr.ewm(com=period - 1, min_periods=period).mean().iloc[-1]), 4)


def find_support_resistance(df, current_price):
    levels = []
    for w in [10, 20, 50]:
        subset = df.tail(max(w * 2, 30))
        for i in range(w, len(subset) - w):
            if subset["Low"].iloc[i] == subset["Low"].iloc[i - w: i + w + 1].min():
                levels.append(round(float(subset["Low"].iloc[i]), 4))
            if subset["High"].iloc[i] == subset["High"].iloc[i - w: i + w + 1].max():
                levels.append(round(float(subset["High"].iloc[i]), 4))

    for ma_col in ["MA20", "MA50", "MA100", "MA200"]:
        if ma_col in df.columns:
            val = float(df[ma_col].iloc[-1])
            if not math.isnan(val):
                levels.append(round(val, 4))

    levels = sorted(set(levels))
    clustered, used = [], set()
    for lv in levels:
        if lv in used:
            continue
        cluster  = [x for x in levels if abs(x - lv) / lv < 0.015]
        clustered.append(round(sum(cluster) / len(cluster), 4))
        for x in cluster:
            used.add(x)

    supports    = sorted([l for l in clustered if l < current_price * 0.99], reverse=True)
    resistances = sorted([l for l in clustered if l > current_price * 1.01])
    return {"supports": supports[:5], "resistances": resistances[:5]}


# ─────────────────────────────────────────────────────────────────────────────
# POSITION SIZING
# ─────────────────────────────────────────────────────────────────────────────

def compute_position_sizing(entry_mid: float, stop_loss: float,
                            capital: float = CAPITAL_EUR,
                            risk_pct: float = RISK_PCT,
                            fx_rate: float = 1.0,
                            adv_20: float = None) -> dict:
    """
    FX-aware fixed-risk position sizing with two hard caps (Improvement 6).

    Cap 1 — ADV: never trade > 10 % of the 20-day average daily volume.
    Cap 2 — Concentration: never put > 20 % of capital into a single position.
    """
    risk_eur      = capital * risk_pct
    risk_local    = risk_eur * fx_rate
    stop_distance = abs(entry_mid - stop_loss)
    if stop_distance <= 0:
        return None

    num_shares = risk_local / stop_distance
    capped_by  = None

    # Cap 1: ADV — max 10 % of 20-day average daily volume
    if adv_20 and adv_20 > 0:
        adv_cap = adv_20 * 0.10
        if num_shares > adv_cap:
            num_shares = adv_cap
            capped_by  = "ADV"

    # Cap 2: max 20 % of capital in one name
    max_pos_local = (capital * 0.20) * fx_rate
    pos_local     = num_shares * entry_mid
    if pos_local > max_pos_local:
        num_shares = max_pos_local / entry_mid
        capped_by  = "MAX_POS" if capped_by is None else f"{capped_by}+MAX_POS"

    num_shares  = max(1, int(round(num_shares)))
    pos_local   = num_shares * entry_mid
    pos_eur     = pos_local / fx_rate
    risk_actual = num_shares * stop_distance
    risk_act_eur= risk_actual / fx_rate

    return {
        "risk_eur":        round(risk_act_eur),
        "position_eur":    round(pos_eur),
        "position_local":  round(pos_local),
        "num_shares":      num_shares,
        "risk_pct_used":   round(risk_pct * 100, 1),
        "sizing_capped_by": capped_by,
    }


# ─────────────────────────────────────────────────────────────────────────────
# CORE TA ENTRY ENGINE
# ─────────────────────────────────────────────────────────────────────────────

def run_ta_entry(ticker, scanner_data=None):
    result = {
        "ticker": ticker, "name": ticker, "status": "error", "error": None,
        "current_price": None, "currency": "USD",
        "entry_low": None, "entry_high": None,
        "stop_loss": None, "stop_pct": None,
        "target_1": None, "rr_t1": None,
        "target_2": None, "rr_t2": None,
        "recommended_exit": None, "rr_exit": None,
        "verdict": None, "rsi": None, "macd": None, "adx": None,
        "trend_primary": None, "trend_secondary": None, "momentum_st": None,
        "ma50": None, "ma200": None,
        "week52_high": None, "week52_low": None,
        "atr": None, "supports": [], "resistances": [],
        "catalyst_note": None, "pass_reason": None,
        # Position sizing
        "position_size_eur": None, "position_size_local": None,
        "risk_eur": None, "num_shares": None, "risk_pct_used": None,
        # Bias + catalyst tier inherited from scanner
        "bias": "LONG", "catalyst_tier": "B",
        # Scanner context carried forward for display
        "rs_vs_bench": None, "day_return": None, "rvol": None,
        # Improvement 1 + 4: news and weekly trend from scanner
        "has_news_today": False, "news_headline": "",
        "weekly_trend_up": None,
        # Improvements 3, 5, 6, 7, 8, 9, 10, 11
        "rs_vs_sector":          None,
        "short_interest_pct":    None,
        "earnings_surprise_avg": None,
        "futures_roll_warning":  False,
        "days_to_roll":          None,
        "roll_date":             None,
        "repeat_days":           0,
        "score_trend":           "→",
        "sizing_capped_by":      None,
        "corr_flag":             [],
    }

    # ── Inherit scanner context early — bias is needed before level calculations
    if scanner_data:
        result["bias"]            = scanner_data.get("bias", "LONG")
        result["catalyst_tier"]   = scanner_data.get("catalyst_tier", "B")
        result["rs_vs_bench"]     = scanner_data.get("rs_vs_bench")
        result["day_return"]      = scanner_data.get("day_return")
        result["rvol"]            = scanner_data.get("rvol")
        result["has_news_today"]        = scanner_data.get("has_news_today", False)
        result["news_headline"]         = scanner_data.get("news_headline", "")
        result["weekly_trend_up"]       = scanner_data.get("weekly_trend_up")
        result["rs_vs_sector"]          = scanner_data.get("rs_vs_sector")
        result["short_interest_pct"]    = scanner_data.get("short_interest_pct")
        result["earnings_surprise_avg"] = scanner_data.get("earnings_surprise_avg")
        result["futures_roll_warning"]  = scanner_data.get("futures_roll_warning", False)
        result["days_to_roll"]          = scanner_data.get("days_to_roll")
        result["roll_date"]             = scanner_data.get("roll_date")
        result["repeat_days"]           = scanner_data.get("repeat_days", 0)
        result["score_trend"]           = scanner_data.get("score_trend", "→")
    bias = result["bias"]

    try:
        end_dt   = datetime.utcnow()
        start_dt = end_dt - timedelta(days=LOOKBACK_DAYS + 50)
        df = yf.download(
            ticker,
            start=start_dt.strftime("%Y-%m-%d"),
            end=end_dt.strftime("%Y-%m-%d"),
            progress=False
        )
        if df.empty or len(df) < 60:
            result["error"] = "Insufficient price history"
            return result

        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)
        df = df.dropna(subset=["Close", "High", "Low", "Volume"])

        current_price         = float(df["Close"].iloc[-1])
        result["current_price"] = round(current_price, 4)

        try:
            info = yf.Ticker(ticker).info
            result["currency"] = info.get("currency", "USD")
            result["name"] = info.get("longName") or info.get("shortName") or ticker
            ed = info.get("earningsDate") or info.get("earningsTimestamp")
            if ed:
                if isinstance(ed, (int, float)):
                    ed = datetime.fromtimestamp(ed).strftime("%Y-%m-%d")
                result["catalyst_note"] = f"Earnings: {ed}"
        except Exception:
            pass

        # Fetch FX rate once per currency (cached) — needed for local-currency sizing
        fx_rate = get_eur_fx_rate(result["currency"])

        df["MA20"]  = df["Close"].rolling(20).mean()
        df["MA50"]  = df["Close"].rolling(50).mean()
        df["MA100"] = df["Close"].rolling(100).mean()
        df["MA200"] = df["Close"].rolling(200).mean()

        ma20  = float(df["MA20"].iloc[-1])
        ma50  = float(df["MA50"].iloc[-1])
        ma200 = float(df["MA200"].iloc[-1])
        result["ma50"]  = round(ma50, 4)
        result["ma200"] = round(ma200, 4)

        last_252 = df.tail(252)
        result["week52_high"] = round(float(last_252["High"].max()), 4)
        result["week52_low"]  = round(float(last_252["Low"].min()), 4)

        rsi             = compute_rsi(df["Close"])
        macd, sig, hist = compute_macd(df["Close"])
        adx             = compute_adx(df["High"], df["Low"], df["Close"])
        atr             = compute_atr(df["High"], df["Low"], df["Close"])
        result["rsi"]   = rsi
        result["macd"]  = macd
        result["adx"]   = adx
        result["atr"]   = round(atr, 4)

        if current_price > ma50 and ma50 > ma200:
            primary = "UPTREND"
        elif current_price < ma50 and ma50 < ma200:
            primary = "DOWNTREND"
        else:
            primary = "MIXED"

        secondary   = "UPTREND" if current_price > ma20 else "DOWNTREND"
        if bias == "SHORT":
            # For SHORT: bearish signals confirm direction (desired)
            momentum_st = "BEARISH" if (rsi < 45 and hist < 0) else ("BULLISH" if (rsi > 55 and hist > 0) else "NEUTRAL")
        else:
            momentum_st = "BULLISH" if (rsi > 55 and hist > 0) else ("BEARISH" if (rsi < 45 and hist < 0) else "NEUTRAL")

        result["trend_primary"]   = primary
        result["trend_secondary"] = secondary
        result["momentum_st"]     = momentum_st

        levels      = find_support_resistance(df, current_price)
        supports    = levels["supports"]
        resistances = levels["resistances"]
        result["supports"]    = supports
        result["resistances"] = resistances

        if bias == "SHORT":
            # ── SHORT trade: sell near resistance, stop above, targets below ──
            nearest_resistance = resistances[0] if resistances else current_price * 1.05
            entry_high = round(nearest_resistance * 0.995, 4)
            entry_low  = round(current_price * 0.995, 4)
            if entry_low > entry_high:
                entry_low, entry_high = entry_high, entry_low
            entry_mid  = round((entry_low + entry_high) / 2, 4)

            stop_loss = round(nearest_resistance + 1.5 * atr, 4)
            stop_pct  = round((stop_loss - entry_mid) / entry_mid, 4)

            if stop_pct < MIN_STOP_PCT:
                stop_loss = round(entry_mid * (1 + MIN_STOP_PCT), 4)
                stop_pct  = MIN_STOP_PCT
            if stop_pct > MAX_STOP_PCT:
                result["status"]      = "pass"
                result["pass_reason"] = f"Short stop {stop_pct:.1%} exceeds 25% max"
                result["verdict"]     = "PASS"
                return result

            result["entry_low"]  = entry_low
            result["entry_high"] = entry_high
            result["stop_loss"]  = stop_loss
            result["stop_pct"]   = round(stop_pct * 100, 2)

            risk = stop_loss - entry_mid   # positive distance from entry to stop

            # Downside targets — nearest support levels below entry (supports sorted descending)
            t1 = next((s for s in supports if s < entry_low * 0.99), None)
            if t1 is None:
                t1 = round(entry_mid - 1.5 * risk, 4)
            rr_t1 = round((entry_mid - t1) / risk, 2)
            if rr_t1 < MIN_RR_T1:
                result["status"]      = "pass"
                result["pass_reason"] = f"Short T1 R/R {rr_t1:.2f} below minimum {MIN_RR_T1}"
                result["verdict"]     = "PASS"
                return result

            t2 = next((s for s in supports if s < t1 * 0.99), None)
            if t2 is None:
                t2 = round(entry_mid - 2.5 * risk, 4)
            rr_t2 = round((entry_mid - t2) / risk, 2)
            if rr_t2 < MIN_RR_T2:
                t2    = round(entry_mid - 2.5 * risk, 4)
                rr_t2 = round((entry_mid - t2) / risk, 2)

            exit_price = next((s for s in supports if s < t2 * 0.98), None)
            if exit_price is None:
                exit_price = round(entry_mid - 4 * risk, 4)
            # Apply 52w low floor only when it still leaves room below T2
            w52_floor = round(result["week52_low"] * 1.05, 4)
            if w52_floor < t2:
                exit_price = max(exit_price, w52_floor)
            exit_price = min(exit_price, round(t2 * 0.98, 4))  # ceiling: always below T2
            rr_exit    = round((entry_mid - exit_price) / risk, 2)

            result["target_1"]         = round(t1, 4)
            result["rr_t1"]            = rr_t1
            result["target_2"]         = round(t2, 4)
            result["rr_t2"]            = rr_t2
            result["recommended_exit"] = round(exit_price, 4)
            result["rr_exit"]          = rr_exit

            adv_20 = float(df["Volume"].tail(20).mean()) if "Volume" in df.columns else None
            sizing = compute_position_sizing(entry_mid, stop_loss, fx_rate=fx_rate, adv_20=adv_20)
            if sizing:
                result["position_size_eur"]   = sizing["position_eur"]
                result["position_size_local"] = sizing["position_local"]
                result["risk_eur"]            = sizing["risk_eur"]
                result["num_shares"]          = sizing["num_shares"]
                result["risk_pct_used"]       = sizing["risk_pct_used"]
                result["sizing_capped_by"]    = sizing["sizing_capped_by"]

            # Verdict for SHORT
            at_support_sh = bool(supports and current_price <= supports[0] * 1.03)
            if at_support_sh:
                verdict = f"WAIT FOR BREAKDOWN BELOW {supports[0]:.2f}"
            elif momentum_st == "BULLISH":
                verdict = f"WAIT — Counter-trend momentum (RSI {rsi:.0f})"
            else:
                verdict = "ENTER SHORT NOW"

        else:
            # ── LONG trade (original logic — unchanged) ───────────────────────
            nearest_support = supports[0] if supports else current_price * 0.95
            entry_low  = round(nearest_support * 1.005, 4)
            entry_high = round(current_price * 1.005, 4)
            entry_mid  = round((entry_low + entry_high) / 2, 4)

            stop_loss = round(nearest_support - 1.5 * atr, 4)
            stop_pct  = round((entry_mid - stop_loss) / entry_mid, 4)

            if stop_pct < MIN_STOP_PCT:
                stop_loss = round(entry_mid * (1 - MIN_STOP_PCT), 4)
                stop_pct  = MIN_STOP_PCT

            if stop_pct > MAX_STOP_PCT:
                result["status"]      = "pass"
                result["pass_reason"] = f"Stop {stop_pct:.1%} exceeds 25% max -- no clean stop"
                result["verdict"]     = "PASS"
                return result

            result["entry_low"]  = entry_low
            result["entry_high"] = entry_high
            result["stop_loss"]  = stop_loss
            result["stop_pct"]   = round(stop_pct * 100, 2)

            risk = entry_mid - stop_loss

            t1    = next((r for r in resistances if r > entry_high), None)
            if t1 is None:
                t1 = round(entry_mid + 1.5 * risk, 4)
            rr_t1 = round((t1 - entry_mid) / risk, 2)
            if rr_t1 < MIN_RR_T1:
                result["status"]      = "pass"
                result["pass_reason"] = f"T1 R/R {rr_t1:.2f} below minimum {MIN_RR_T1}"
                result["verdict"]     = "PASS"
                return result

            t2    = next((r for r in resistances if r > t1 * 1.01), None)
            if t2 is None:
                t2 = round(entry_mid + 2.5 * risk, 4)
            rr_t2 = round((t2 - entry_mid) / risk, 2)
            if rr_t2 < MIN_RR_T2:
                t2    = round(entry_mid + 2.5 * risk, 4)
                rr_t2 = round((t2 - entry_mid) / risk, 2)

            exit_price = next((r for r in resistances if r > t2 * 1.02), None)
            if exit_price is None:
                exit_price = round(entry_mid + 4 * risk, 4)
            # Apply 52w high cap only when it still leaves room above T2
            w52_cap = round(result["week52_high"] * 0.95, 4)
            if w52_cap > t2:
                exit_price = min(exit_price, w52_cap)
            exit_price = max(exit_price, round(t2 * 1.02, 4))  # floor: always above T2
            rr_exit    = round((exit_price - entry_mid) / risk, 2)

            result["target_1"]         = round(t1, 4)
            result["rr_t1"]            = rr_t1
            result["target_2"]         = round(t2, 4)
            result["rr_t2"]            = rr_t2
            result["recommended_exit"] = round(exit_price, 4)
            result["rr_exit"]          = rr_exit

            adv_20 = float(df["Volume"].tail(20).mean()) if "Volume" in df.columns else None
            sizing = compute_position_sizing(entry_mid, stop_loss, fx_rate=fx_rate, adv_20=adv_20)
            if sizing:
                result["position_size_eur"]   = sizing["position_eur"]
                result["position_size_local"] = sizing["position_local"]
                result["risk_eur"]            = sizing["risk_eur"]
                result["num_shares"]          = sizing["num_shares"]
                result["risk_pct_used"]       = sizing["risk_pct_used"]
                result["sizing_capped_by"]    = sizing["sizing_capped_by"]

            at_support      = current_price <= nearest_support * 1.02
            near_resistance = bool(resistances and current_price >= resistances[0] * 0.97)
            at_52w_low      = current_price <= result["week52_low"] * 1.05

            if at_support and momentum_st != "BEARISH" and primary != "DOWNTREND":
                verdict = "ENTER NOW"
            elif current_price <= entry_high and primary == "UPTREND":
                verdict = "ENTER NOW"
            elif near_resistance:
                verdict = f"WAIT FOR BREAKOUT ABOVE {resistances[0]:.2f}"
            elif current_price > entry_high:
                verdict = f"WAIT FOR DIP TO {entry_low:.2f}"
            else:
                verdict = "ENTER NOW"

            if primary == "DOWNTREND" and not at_52w_low:
                verdict = f"WAIT FOR BREAKOUT ABOVE {ma50:.2f} (50-day MA)"

        result["verdict"] = verdict
        result["status"]  = "ok"

    except Exception as e:
        result["error"] = str(e)

    return result


# ─────────────────────────────────────────────────────────────────────────────
# RANKING — highest return probability first
# ─────────────────────────────────────────────────────────────────────────────

def rank_by_return_probability(actionable, top_n=TOP_N):
    """
    Rank: tier first (A+ > A > B), then entry quality within each tier.
    Direction-aware: DOWNTREND is good alignment for SHORT, UPTREND for LONG.

    Score components (within-tier, max ~100):
      Verdict urgency  — ENTER NOW=30, WAIT FOR DIP/BREAKDOWN=20, else=10
      R/R quality      — rr_exit up to 6:1 mapped to 0-30 pts
      Trend alignment  — direction-aware: 20 / 10 / 5
      RSI              — direction-aware: optimal zone=20, edge=15, poor=5
    """
    tier_priority = {"A+": 300, "A": 200, "B": 100}

    def score(r):
        tier  = tier_priority.get(r.get("catalyst_tier", "B"), 100)
        bias  = r.get("bias", "LONG")
        v     = r.get("verdict", "")
        vs    = 30 if v.startswith("ENTER") else (20 if ("DIP" in v or "BREAKDOWN" in v) else 10)
        rr    = min(r.get("rr_exit", 0) / 6.0, 1.0) * 30
        trend = r.get("trend_primary", "")
        if bias == "SHORT":
            ts = 20 if trend == "DOWNTREND" else (10 if trend == "MIXED" else 5)
        else:
            ts = 20 if trend == "UPTREND" else (10 if trend == "MIXED" else 5)
        rsi = r.get("rsi") or 50
        if bias == "SHORT":
            rs = 20 if rsi <= 45 else (15 if rsi <= 55 else 5)
        else:
            rs = 20 if 40 <= rsi <= 65 else (15 if 65 < rsi <= 75 else 5)
        # Improvement 8: penalise stale fading signals
        # A ticker qualifying for 5+ consecutive days with declining score = momentum exhausted
        repeat = r.get("repeat_days", 0)
        trend_s = r.get("score_trend", "→")
        staleness = -15 if (repeat >= 5 and trend_s == "↓") else (-5 if repeat >= 3 and trend_s == "↓" else 0)
        return tier + vs + rr + ts + rs + staleness

    return sorted(actionable, key=score, reverse=True)[:top_n]


# ─────────────────────────────────────────────────────────────────────────────
# OUTPUT FORMATTERS
# ─────────────────────────────────────────────────────────────────────────────

def compute_correlation_flags(top10: list) -> None:
    """
    Improvement 11 — Correlation / portfolio overlap guard.
    Compute 60-day return correlations for all top10 tickers.
    Sets corr_flag = list of ticker strings with |corr| >= 0.75.
    Modifies top10 in-place.
    """
    tickers = [r["ticker"] for r in top10]
    for r in top10:
        r["corr_flag"] = []
    if len(tickers) < 2:
        return
    try:
        end_dt   = datetime.utcnow()
        start_dt = end_dt - timedelta(days=90)
        raw = yf.download(
            tickers,
            start=start_dt.strftime("%Y-%m-%d"),
            end=end_dt.strftime("%Y-%m-%d"),
            progress=False, auto_adjust=True,
            group_by="ticker"
        )
        if raw is None or raw.empty:
            return
        # Build closes DataFrame
        closes = {}
        for t in tickers:
            try:
                if isinstance(raw.columns, __import__("pandas").MultiIndex):
                    col_data = raw[t]["Close"] if t in raw.columns.get_level_values(0) else None
                else:
                    col_data = raw["Close"] if len(tickers) == 1 else None
                if col_data is not None and len(col_data.dropna()) > 20:
                    closes[t] = col_data.astype(float)
            except Exception:
                pass
        if len(closes) < 2:
            return
        df_closes = __import__("pandas").DataFrame(closes).dropna()
        rets      = df_closes.pct_change().dropna()
        corr      = rets.corr()
        for r in top10:
            t = r["ticker"]
            if t not in corr.columns:
                continue
            flagged = [
                other for other in corr.columns
                if other != t and abs(corr.loc[t, other]) >= 0.75
            ]
            r["corr_flag"] = flagged
    except Exception:
        pass


def format_best_opportunities_summary(top10: list, total_actionable: int, total_scanned: int) -> str:
    now  = datetime.utcnow().strftime("%H:%M UTC")
    lines = [
        f"🎯 *BEST OPPORTUNITIES — {now}*",
        f"Ranked by entry quality | {total_actionable} actionable from {total_scanned} scanned",
        f"",
    ]
    trend_tag   = lambda t: "MA↑" if t == "UPTREND" else ("MA→" if t == "MIXED" else "MA↓")
    verdict_tag = lambda v: ("🔴 SHORT" if v.startswith("ENTER SHORT") else
                             ("✅ ENTER"  if v.startswith("ENTER") else
                              ("⏳ BRKDWN" if "BREAKDOWN" in v else
                               ("⏳ DIP"    if "DIP" in v else "⏳ WAIT"))))
    ct_tag  = lambda c: {"A+": "🔥A+", "A": "🔵A", "B": "🟡B"}.get(c, "🟡B")
    bias_tag = lambda b: "📉 SHORT" if b == "SHORT" else "📈 LONG"
    for i, r in enumerate(top10, 1):
        sym    = r.get("ticker", "?")
        trend  = trend_tag(r.get("trend_primary", ""))
        rsi    = r.get("rsi", "-")
        tag    = verdict_tag(r.get("verdict", ""))
        ct     = ct_tag(r.get("catalyst_tier", "B"))
        bias   = bias_tag(r.get("bias", "LONG"))
        day_r  = r.get("day_return")
        rvol_v = r.get("rvol")
        move   = f" {'🟢' if day_r and day_r > 0 else '🔴'}{day_r:+.1f}%" if day_r is not None else ""
        vol    = f" RVOL {rvol_v:.1f}x" if rvol_v is not None else ""
        lines.append(f"{i}. {ct} *{sym}*{move}{vol} | {bias} | {trend} | RSI {rsi} | {tag}")
    lines += ["", "_Detail cards follow_ ↓"]
    return "\n".join(lines)


def format_telegram_card(r, rank):
    ccy      = r.get("currency", "USD")
    sym      = r.get("ticker", "?")
    name     = r.get("name", sym)
    price    = r.get("current_price", 0)
    mid      = (r["entry_low"] + r["entry_high"]) / 2
    entry    = f"{r['entry_low']:.2f} - {r['entry_high']:.2f}"
    verdict  = r.get("verdict", "-")
    bias     = r.get("bias", "LONG")

    # ── Direction-aware stop and target display ────────────────────────────────
    if bias == "SHORT":
        # Stop is ABOVE entry — loss if price rises past it
        stop   = f"{r['stop_loss']:.2f}  (+{r['stop_pct']:.1f}% ▲ above entry)"
        # Targets are BELOW entry — gain when price falls
        t1_pct = (mid - r["target_1"]) / mid * 100
        t2_pct = (mid - r["target_2"]) / mid * 100
        ex_pct = (mid - r["recommended_exit"]) / mid * 100
        t1     = f"{r['target_1']:.2f}  (-{t1_pct:.1f}% ▼)  R/R {r['rr_t1']:.1f}:1"
        t2     = f"{r['target_2']:.2f}  (-{t2_pct:.1f}% ▼)  R/R {r['rr_t2']:.1f}:1"
        ex     = f"{r['recommended_exit']:.2f}  (-{ex_pct:.1f}% ▼)  R/R {r['rr_exit']:.1f}:1"
    else:
        # Stop is BELOW entry — loss if price falls past it
        stop   = f"{r['stop_loss']:.2f}  (-{r['stop_pct']:.1f}%)"
        t1_pct = (r["target_1"] - mid) / mid * 100
        t2_pct = (r["target_2"] - mid) / mid * 100
        ex_pct = (r["recommended_exit"] - mid) / mid * 100
        t1     = f"{r['target_1']:.2f}  (+{t1_pct:.1f}%)  R/R {r['rr_t1']:.1f}:1"
        t2     = f"{r['target_2']:.2f}  (+{t2_pct:.1f}%)  R/R {r['rr_t2']:.1f}:1"
        ex     = f"{r['recommended_exit']:.2f}  (+{ex_pct:.1f}%)  R/R {r['rr_exit']:.1f}:1"

    # ── Direction-aware verdict tag ────────────────────────────────────────────
    if verdict.startswith("ENTER SHORT"):
        tag = "🔴 ENTER SHORT NOW"
    elif verdict.startswith("ENTER"):
        tag = "✅ ENTER NOW"
    elif "BREAKDOWN" in verdict:
        tag = f"⏳ {verdict}"
    elif "DIP" in verdict:
        tag = f"⏳ WAIT — {verdict}"
    else:
        tag = f"⏳ {verdict}"

    # ── Direction-aware momentum label ─────────────────────────────────────────
    mom = r.get("momentum_st", "-")
    if bias == "SHORT":
        if mom == "BEARISH":
            momentum_display = "BEARISH ✅ (short confirmed — enter)"
        elif mom == "BULLISH":
            momentum_display = "BULLISH ⚠️ (counter-trend — wait for turn)"
        else:
            momentum_display = "NEUTRAL  (acceptable — proceed with caution)"
    else:
        if mom == "BULLISH":
            momentum_display = "BULLISH ✅ (enter now)"
        elif mom == "BEARISH":
            momentum_display = "BEARISH ⚠️ (wait for reversal)"
        else:
            momentum_display = "NEUTRAL  (acceptable — watch for confirmation)"

    trend_tag  = {"UPTREND": "MA↑", "MIXED": "MA→", "DOWNTREND": "MA↓"}.get(r.get("trend_primary", ""), "MA→")
    ct         = r.get("catalyst_tier", "B")
    ct_label   = {"A+": "🔥A+", "A": "🔵A", "B": "🟡B"}.get(ct, "🟡B")
    bias_label = "📉 SHORT" if bias == "SHORT" else "📈 LONG"
    catalyst   = r.get("catalyst_note") or "No near-term catalyst"

    # ── Improvement 1: News headline ──────────────────────────────────────────
    has_news     = r.get("has_news_today", False)
    news_headline = r.get("news_headline", "")
    if has_news and news_headline:
        short_hl  = (news_headline[:80] + "…") if len(news_headline) > 82 else news_headline
        news_line = f"📰 {short_hl}"
    else:
        news_line = None

    # ── Improvement 4: Weekly trend label ─────────────────────────────────────
    wtu = r.get("weekly_trend_up")
    if wtu is True:
        weekly_label = "W↑ (above 20-wk MA)"
    elif wtu is False:
        weekly_label = "W↓ (below 20-wk MA)"
    else:
        weekly_label = None

    # ── Scanner context line (today's move, RVOL, RS vs market) ───────────────
    day_r  = r.get("day_return")
    rvol_v = r.get("rvol")
    rs_v   = r.get("rs_vs_bench")
    ctx    = []
    if day_r is not None:
        ctx.append(f"{'🟢' if day_r > 0 else '🔴'}{day_r:+.1f}% today")
    if rvol_v is not None:
        ctx.append(f"RVOL {rvol_v:.1f}x")
    if rs_v is not None:
        ctx.append(f"RS {rs_v:+.1f}% vs mkt")
    context_line = "  |  ".join(ctx) if ctx else None

    # ── Position sizing ────────────────────────────────────────────────────────
    pos_eur   = r.get("position_size_eur")
    pos_local = r.get("position_size_local")
    risk_eur  = r.get("risk_eur")
    n_shares  = r.get("num_shares")
    risk_pct  = r.get("risk_pct_used", RISK_PCT * 100)
    if pos_local is not None and n_shares is not None:
        if ccy == "EUR":
            sizing_line = f"Size:     €{pos_local:,.0f}  ({n_shares:,} units)  Risk: €{risk_eur:,.0f} ({risk_pct}%)"
        else:
            sizing_line = (f"Size:     {ccy} {pos_local:,.0f}  ({n_shares:,} units)"
                           f"  ≈ €{pos_eur:,.0f}  Risk: €{risk_eur:,.0f} ({risk_pct}%)")
    else:
        sizing_line = "Size:     —"

    card_lines = [
        f"*#{rank} {ct_label} {sym}* ({name}) — {bias_label}",
    ]
    if context_line:
        card_lines.append(context_line)
    if weekly_label:
        card_lines.append(f"Weekly:   {weekly_label}")
    card_lines += [
        f"{ccy} {price:.2f} | {trend_tag} | RSI {r.get('rsi', '-')}",
        f"",
        f"Entry:    {ccy} {entry}",
        f"Stop:     {ccy} {stop}",
        f"T1:       {ccy} {t1}",
        f"T2:       {ccy} {t2}",
        f"Exit:     {ccy} {ex}",
        f"",
        f"{sizing_line}",
        f"Momentum: {momentum_display}",
        f"Catalyst: {catalyst}",
    ]
    if news_line:
        card_lines.append(news_line)

    # Improvement 9: Short interest
    si_pct = r.get("short_interest_pct")
    if si_pct is not None:
        si_flag = " 🔥 Squeeze watch" if si_pct >= 15 else ""
        card_lines.append(f"Short int: {si_pct:.0f}%{si_flag}")

    # Improvement 10: Earnings surprise history
    earn_surp = r.get("earnings_surprise_avg")
    if earn_surp is not None:
        sign = "+" if earn_surp >= 0 else ""
        card_lines.append(f"Avg EPS beat: {sign}{earn_surp:.0f}%")

    # Improvement 7: Futures roll warning
    if r.get("futures_roll_warning"):
        dtroll = r.get("days_to_roll")
        card_lines.append(f"⚠️ Futures roll in {dtroll}d — consider sizing down")

    # Improvement 11: Correlation flag
    corr_f = r.get("corr_flag", [])
    if corr_f:
        card_lines.append(f"⚠️ Correlated ≥0.75 with: {', '.join(corr_f)}")

    card_lines += [
        f"{tag}",
        f"─────────────────────────────",
    ]
    return "\n".join(card_lines)


def format_markdown_row(r):
    if r["status"] == "pass":
        return f"| {r['ticker']} | PASS | - | - | - | - | {r.get('pass_reason','-')} |"
    if r["status"] == "error":
        return f"| {r['ticker']} | ERROR | - | - | - | - | {r.get('error','-')} |"
    return (
        f"| {r['ticker']} "
        f"| {r['current_price']:.2f} "
        f"| {r['entry_low']:.2f}-{r['entry_high']:.2f} "
        f"| {r['stop_loss']:.2f} (-{r['stop_pct']:.1f}%) "
        f"| {r['target_1']:.2f} ({r['rr_t1']:.1f}:1) "
        f"| {r['recommended_exit']:.2f} ({r['rr_exit']:.1f}:1) "
        f"| {r['verdict']} |"
    )


# ─────────────────────────────────────────────────────────────────────────────
# TELEGRAM SENDER
# ─────────────────────────────────────────────────────────────────────────────

def send_telegram(text):
    if DRY_RUN:
        print(f"[DRY RUN] Telegram:\n{text}\n")
        return True
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        print("WARN: Telegram credentials not set -- skipping")
        return False
    import urllib.request
    url     = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    payload = json.dumps({
        "chat_id": TELEGRAM_CHAT_ID,
        "text": text,
        "parse_mode": "Markdown",
        "disable_web_page_preview": True
    }).encode()
    try:
        req = urllib.request.Request(
            url, data=payload,
            headers={"Content-Type": "application/json"}
        )
        urllib.request.urlopen(req, timeout=15)
        return True
    except Exception as e:
        print(f"ERROR sending Telegram: {e}")
        return False


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL REPORT
# ─────────────────────────────────────────────────────────────────────────────

def generate_excel_report(all_candidates: list, top10: list,
                          filepath: str = "scan_report.xlsx") -> str:
    """
    Build a 3-sheet Excel workbook:
      Sheet 1 — TOP MOVERS    : all scanner candidates, sorted by score
      Sheet 2 — TOP OPPORTUNITIES : ranked top-10 from TA runner
      Sheet 3 — DETAIL CARDS  : full TA analysis (entry / stop / T1 / T2 / exit / sizing)
    Returns the filepath written.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    # ── Colour palette ────────────────────────────────────────────────────────
    HDR_FILL  = PatternFill("solid", fgColor="1F3864")   # dark navy header
    HDR_FONT  = Font(color="FFFFFF", bold=True, size=10)
    AP_FILL   = PatternFill("solid", fgColor="C55A11")   # burnt orange  A+
    A_FILL    = PatternFill("solid", fgColor="2E75B6")   # blue          A
    B_FILL    = PatternFill("solid", fgColor="FFC000")   # amber         B
    LONG_CLR  = Font(color="00B050", bold=True)          # green  LONG
    SHORT_CLR = Font(color="C00000", bold=True)          # red    SHORT
    OK_FILL   = PatternFill("solid", fgColor="E2EFDA")   # light green row (ENTER)
    ALT_FILL  = PatternFill("solid", fgColor="F2F2F2")   # light grey alternate rows

    def hdr(ws, headers):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = HDR_FILL
            c.font = HDR_FONT
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 30

    def auto_width(ws, headers, min_w=12, max_w=28):
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = min(
                max_w, max(min_w, len(headers[col - 1]) + 2))

    def tier_colour(cell, ct):
        if ct == "A+":
            cell.fill = AP_FILL
            cell.font = Font(bold=True, color="FFFFFF")
        elif ct == "A":
            cell.fill = A_FILL
            cell.font = Font(bold=True, color="FFFFFF")
        else:
            cell.fill = B_FILL
            cell.font = Font(bold=True)

    def bias_colour(cell, bias):
        cell.font = SHORT_CLR if bias == "SHORT" else LONG_CLR

    wb = Workbook()

    # ── Sheet 1 : TOP MOVERS ─────────────────────────────────────────────────
    ws1       = wb.active
    ws1.title = "TOP MOVERS"
    h1 = ["Rank", "Ticker", "Name", "Type", "Catalyst Tier", "Bias",
          "Move %", "RS vs Mkt %", "MA Align", "ATR %", "RVOL", "Score",
          "Earnings", "Repeat Days", "Trend"]
    hdr(ws1, h1)

    # Build a name lookup from top10 (ta_runner fetches names from yfinance)
    name_map = {r.get("ticker"): r.get("name", r.get("ticker", "")) for r in top10}

    sorted_cands = sorted(all_candidates, key=lambda x: x.get("score", 0), reverse=True)
    for ri, c in enumerate(sorted_cands, 2):
        tier_lbl = {"etf": "ETF", "future": "FUT", "stock": "STK"}.get(c.get("tier", "stock"), "STK")
        ct   = c.get("catalyst_tier", "B")
        bias = c.get("bias", "LONG")
        row  = [ri - 1,
                c.get("ticker", ""),
                name_map.get(c.get("ticker", ""), c.get("ticker", "")),
                tier_lbl, ct, bias,
                round(c.get("day_return", 0) or 0, 2),
                round(c.get("rs_vs_bench", 0) or 0, 2),
                c.get("ma_align", ""),
                round(c.get("atr_pct", 0) or 0, 2),
                round(c.get("rvol", 0) or 0, 2),
                round(c.get("score", 0) or 0, 0),
                "Yes" if c.get("earnings_soon") else "No",
                c.get("repeat_days", 0),
                c.get("score_trend", "")]
        bg = ALT_FILL if ri % 2 == 0 else None
        for col, val in enumerate(row, 1):
            cell = ws1.cell(row=ri, column=col, value=val)
            cell.alignment = Alignment(horizontal="center")
            if bg:
                cell.fill = bg
        tier_colour(ws1.cell(row=ri, column=5), ct)
        bias_colour(ws1.cell(row=ri, column=6), bias)

    ws1.freeze_panes = "A2"
    auto_width(ws1, h1)

    # ── Sheet 2 : TOP OPPORTUNITIES ──────────────────────────────────────────
    ws2       = wb.create_sheet("TOP OPPORTUNITIES")
    h2 = ["Rank", "Ticker", "Name", "Catalyst Tier", "Bias",
          "Today Move %", "RVOL", "MA Align", "RSI", "Trend", "Verdict"]
    hdr(ws2, h2)

    trend_lbl = {"UPTREND": "MA↑", "MIXED": "MA→", "DOWNTREND": "MA↓"}
    for ri, r in enumerate(top10, 2):
        ct   = r.get("catalyst_tier", "B")
        bias = r.get("bias", "LONG")
        verdict = r.get("verdict", "")
        row = [ri - 1,
               r.get("ticker", ""),
               r.get("name", r.get("ticker", "")),
               ct, bias,
               round(r.get("day_return") or 0, 2),
               round(r.get("rvol") or 0, 2),
               trend_lbl.get(r.get("trend_primary", ""), "MA→"),
               r.get("rsi", ""),
               r.get("trend_primary", ""),
               verdict]
        row_fill = OK_FILL if verdict.startswith("ENTER") else (ALT_FILL if ri % 2 == 0 else None)
        for col, val in enumerate(row, 1):
            cell = ws2.cell(row=ri, column=col, value=val)
            cell.alignment = Alignment(horizontal="center")
            if row_fill:
                cell.fill = row_fill
        tier_colour(ws2.cell(row=ri, column=4), ct)
        bias_colour(ws2.cell(row=ri, column=5), bias)

    ws2.freeze_panes = "A2"
    auto_width(ws2, h2)

    # ── Sheet 3 : DETAIL CARDS ───────────────────────────────────────────────
    ws3       = wb.create_sheet("DETAIL CARDS")
    h3 = ["Rank", "Ticker", "Name", "Catalyst Tier", "Bias", "Currency",
          "Current Price", "Entry Low", "Entry High",
          "Stop Loss", "Stop %",
          "T1 Price", "T1 R/R",
          "T2 Price", "T2 R/R",
          "Exit Price", "Exit R/R",
          "Position Size (local ccy)", "Units", "Risk €", "Risk %",
          "RSI", "MA Trend", "Momentum", "Catalyst", "Verdict"]
    hdr(ws3, h3)

    for ri, r in enumerate(top10, 2):
        ct      = r.get("catalyst_tier", "B")
        bias    = r.get("bias", "LONG")
        verdict = r.get("verdict", "")
        row = [ri - 1,
               r.get("ticker", ""),
               r.get("name", r.get("ticker", "")),
               ct, bias,
               r.get("currency", "USD"),
               r.get("current_price"),
               r.get("entry_low"),
               r.get("entry_high"),
               r.get("stop_loss"),
               r.get("stop_pct"),
               r.get("target_1"),
               r.get("rr_t1"),
               r.get("target_2"),
               r.get("rr_t2"),
               r.get("recommended_exit"),
               r.get("rr_exit"),
               r.get("position_size_local"),
               r.get("num_shares"),
               r.get("risk_eur"),
               r.get("risk_pct_used"),
               r.get("rsi"),
               trend_lbl.get(r.get("trend_primary", ""), "MA→"),
               r.get("momentum_st", ""),
               r.get("catalyst_note") or "No near-term catalyst",
               verdict]
        row_fill = OK_FILL if verdict.startswith("ENTER") else (ALT_FILL if ri % 2 == 0 else None)
        for col, val in enumerate(row, 1):
            cell = ws3.cell(row=ri, column=col, value=val)
            cell.alignment = Alignment(horizontal="center")
            if row_fill:
                cell.fill = row_fill
            # Number formats
            if col in (7, 8, 9, 10, 12, 14, 16):   # prices
                cell.number_format = "#,##0.00"
            elif col in (11, 13, 15, 17, 21):       # %  and R/R
                cell.number_format = "0.00"
            elif col in (18, 20):                   # euro amounts
                cell.number_format = "#,##0"
            elif col == 19:                         # units
                cell.number_format = "#,##0"
        tier_colour(ws3.cell(row=ri, column=4), ct)
        bias_colour(ws3.cell(row=ri, column=5), bias)
        # Green/red verdict cell
        vc = ws3.cell(row=ri, column=26)
        if verdict.startswith("ENTER SHORT"):
            vc.fill = PatternFill("solid", fgColor="C00000")
            vc.font = Font(bold=True, color="FFFFFF")
        elif verdict.startswith("ENTER"):
            vc.fill = PatternFill("solid", fgColor="00B050")
            vc.font = Font(bold=True, color="FFFFFF")

    ws3.freeze_panes = "A2"
    auto_width(ws3, h3, min_w=13, max_w=22)

    wb.save(filepath)
    return filepath


# ─────────────────────────────────────────────────────────────────────────────
# EMAIL REPORT  (HTML)
# ─────────────────────────────────────────────────────────────────────────────

_CCY_SYM = {
    "USD": "$",  "EUR": "€",  "GBP": "£",  "JPY": "¥",  "KRW": "₩",
    "HKD": "HK$","NOK": "NOK ","CAD": "C$","AUD": "A$", "SEK": "SEK ",
    "DKK": "DKK ","CHF": "CHF ","INR": "₹","SGD": "S$",
}

def _fp(val) -> str:
    """Format price for email table."""
    if val is None:
        return "—"
    try:
        v = float(val)
        if v >= 100000: return f"{v:,.0f}"
        if v >= 10000:  return f"{v:,.0f}"
        if v >= 100:    return f"{v:,.1f}"
        return f"{v:,.2f}"
    except (TypeError, ValueError):
        return "—"

def _sv(v: str) -> str:
    """Shorten verdict to table label."""
    if not v: return "—"
    if v.startswith("ENTER SHORT"): return "Short Now"
    if v.startswith("ENTER"):       return "Enter Now"
    m = re.search(r"DIP TO ([\d,.]+)", v)
    if m: return f"Dip {m.group(1).replace(',','').rstrip('0').rstrip('.')}"
    m = re.search(r"BREAKDOWN BELOW ([\d,.]+)", v)
    if m: return f"BrkDwn {m.group(1).replace(',','')}"
    m = re.search(r"BREAKOUT ABOVE ([\d,.]+)", v)
    if m: return f"Brkout {m.group(1).replace(',','')}"
    if "PASS" in v.upper(): return "Pass"
    return v[:12]

def build_html_email(top10: list, run_time: str) -> str:
    """Build complete HTML email body from top10 TA results."""

    NAVY   = "#1F3864"
    LGRAY  = "#F5F5F5"
    DGRAY  = "#D0D0D0"

    TH = (
        'style="background:#1F3864;color:#FFFFFF;font-weight:bold;'
        'padding:8px 10px;text-align:center;border:1px solid #3A5080;'
        'white-space:nowrap;font-size:11px;"'
    )

    def td(content, bg, extra="", align="center"):
        return (
            f'<td style="padding:7px 9px;border:1px solid #CCCCCC;background:{bg};'
            f'text-align:{align};vertical-align:middle;font-size:12px;'
            f'font-family:Arial,sans-serif;{extra}">{content}</td>'
        )

    def tier_badge(ct):
        cfg = {
            "A+": ("#C55A11", "#FFFFFF", "🔥 A+"),
            "A":  ("#2E75B6", "#FFFFFF", "🔵 A"),
            "B":  ("#FFC000", "#7F6000", "🟡 B"),
        }.get(ct, ("#FFC000", "#7F6000", "🟡 B"))
        return (
            f'<span style="background:{cfg[0]};color:{cfg[1]};font-weight:bold;'
            f'padding:2px 8px;border-radius:4px;font-size:11px;white-space:nowrap;">'
            f'{cfg[2]}</span>'
        )

    def bias_html(bias):
        if bias == "SHORT":
            return '<span style="color:#C00000;font-weight:bold;">📉 SHORT</span>'
        return '<span style="color:#00B050;font-weight:bold;">📈 LONG</span>'

    def momentum_html(mom, bias):
        # For SHORT: bearish is good; for LONG: bullish is good
        if bias == "SHORT":
            if mom == "BEARISH": return '✅ BEARISH'
            if mom == "BULLISH": return '⚠️ BULLISH'
        else:
            if mom == "BULLISH": return '✅ BULLISH'
            if mom == "BEARISH": return '⚠️ BEARISH'
        return '— NEUTRAL'

    trend_map = {"UPTREND": "MA↑", "DOWNTREND": "MA↓", "MIXED": "MA→"}

    cols = ["#", "TICKER", "TIER", "BIAS", "NOW", "ENTRY",
            "STOP", "T1", "T2", "EXIT", "R/R",
            "MOMENTUM", "DAILY TREND", "WEEKLY", "CATALYST", "VERDICT"]
    header = "".join(f"<th {TH}>{c}</th>" for c in cols)

    rows = []
    for i, r in enumerate(top10):
        bias    = r.get("bias", "LONG")
        verdict = r.get("verdict", "")
        bg      = "#EAF5EA" if verdict.startswith("ENTER") else ("#FFF8E1" if "WAIT" in verdict else ("#F5F5F5" if i % 2 == 0 else "#FFFFFF"))

        ccy = r.get("currency", "USD")
        sym = _CCY_SYM.get(ccy, f"{ccy} ")

        name           = r.get("name", r.get("ticker", ""))
        ticker         = r.get("ticker", "")
        has_news       = r.get("has_news_today", False)
        news_headline  = r.get("news_headline", "")
        name_disp      = (name[:20] + "…") if len(name) > 22 else name
        news_badge   = ' <span title="News today" style="color:#C55A11;font-size:11px;">📰</span>' if has_news else ""
        roll_warn    = r.get("futures_roll_warning", False)
        days_to_roll = r.get("days_to_roll")
        roll_badge   = (f' <span title="Futures roll in {days_to_roll}d" '
                        f'style="color:#C00000;font-size:10px;font-weight:bold;">⚠️ Roll {days_to_roll}d</span>'
                        if roll_warn and days_to_roll is not None else "")
        repeat       = r.get("repeat_days", 0)
        strend       = r.get("score_trend", "→")
        repeat_badge = (f'<br><span style="font-size:10px;color:#888888;">🔄 {repeat}d {strend}</span>'
                        if repeat >= 2 else "")
        rs_b         = r.get("rs_vs_bench")
        rs_s         = r.get("rs_vs_sector")
        corr_f       = r.get("corr_flag", [])
        rs_line      = ""
        if rs_b is not None or rs_s is not None:
            parts = []
            if rs_b is not None:
                col = "#00B050" if rs_b >= 0 else "#C00000"
                parts.append(f'<span style="color:{col};">RS idx: {rs_b:+.1f}%</span>')
            if rs_s is not None:
                col = "#00B050" if rs_s >= 0 else "#C00000"
                parts.append(f'<span style="color:{col};">RS sec: {rs_s:+.1f}%</span>')
            rs_line = f'<br><span style="font-size:10px;">{" | ".join(parts)}</span>'
        corr_others  = ", ".join(corr_f[:2])
        corr_title   = ", ".join(corr_f)
        corr_line    = (
            f'<br><span style="font-size:10px;color:#7030A0;" '
            f'title="Correlated ≥0.75 with {corr_title}">⚠️ corr: {corr_others}</span>'
            if corr_f else ""
        )
        ticker_cell  = (
            f'<strong style="font-size:13px;">{ticker}{news_badge}{roll_badge}</strong>'
            f'<br><span style="font-size:10px;color:#666666;">{name_disp}</span>'
            f'{repeat_badge}{rs_line}{corr_line}'
        )

        price  = f"{sym}{_fp(r.get('current_price'))}"
        entry  = f"{_fp(r.get('entry_low'))}–{_fp(r.get('entry_high'))}"
        stop   = _fp(r.get("stop_loss"))
        t1     = _fp(r.get("target_1"))
        t2     = _fp(r.get("target_2"))
        ex     = _fp(r.get("recommended_exit"))
        rr_val = r.get("rr_exit")
        rr     = f"{rr_val:.1f}×" if rr_val else "—"
        mom    = r.get("momentum_st", "NEUTRAL")
        trend  = trend_map.get(r.get("trend_primary", ""), "MA→")
        ct     = r.get("catalyst_tier", "B")

        # ── Weekly trend cell ──
        wtu = r.get("weekly_trend_up")
        if wtu is True:
            weekly_html = '<span style="color:#00B050;font-weight:bold;font-size:14px;" title="Price above 20-week MA">▲ W↑</span>'
        elif wtu is False:
            weekly_html = '<span style="color:#C00000;font-weight:bold;font-size:14px;" title="Price below 20-week MA">▼ W↓</span>'
        else:
            weekly_html = '<span style="color:#888888;">—</span>'

        # ── Catalyst: show SI, earnings surprise, news headline ──
        cat_note    = r.get("catalyst_note") or "None"
        si_pct      = r.get("short_interest_pct")
        earn_surp   = r.get("earnings_surprise_avg")
        cat_extras  = []
        if si_pct is not None:
            si_col = "#C00000" if si_pct >= 15 else "#888888"
            cat_extras.append(
                f'<span style="color:{si_col};font-size:10px;">SI: {si_pct:.0f}%</span>'
            )
        if earn_surp is not None:
            es_col = "#00B050" if earn_surp >= 0 else "#C00000"
            cat_extras.append(
                f'<span style="color:{es_col};font-size:10px;">Avg beat: {earn_surp:+.0f}%</span>'
            )
        if has_news and news_headline:
            short_hl = (news_headline[:55] + "…") if len(news_headline) > 57 else news_headline
            cat_extras.append(
                f'<span style="color:#C55A11;font-size:10px;font-style:italic;">📰 {short_hl}</span>'
            )
        cat_html = cat_note
        if cat_extras:
            cat_html += "<br>" + "<br>".join(cat_extras)

        rows.append(
            "<tr>"
            + td(str(i + 1),               bg)
            + td(ticker_cell,              bg, align="left")
            + td(tier_badge(ct),           bg)
            + td(bias_html(bias),          bg)
            + td(price,                    bg, "font-weight:bold;")
            + td(entry,                    bg)
            + td(stop,                     bg)
            + td(t1,                       bg)
            + td(t2,                       bg)
            + td(ex,                       bg)
            + td(rr,                       bg, "font-weight:bold;")
            + td(momentum_html(mom, bias), bg, "font-size:11px;")
            + td(trend,                    bg, "font-size:14px;font-weight:bold;")
            + td(weekly_html,              bg, "text-align:center;")
            + td(cat_html,                 bg, "font-size:11px;", align="left")
            + td(_sv(verdict),             bg, "font-size:11px;font-weight:bold;")
            + "</tr>"
        )

    table = (
        f'<table style="border-collapse:collapse;width:100%;font-family:Arial,sans-serif;">'
        f"<thead><tr>{header}</tr></thead>"
        f"<tbody>{''.join(rows)}</tbody>"
        f"</table>"
    )

    legend = f"""
<table style="border-collapse:collapse;width:100%;font-size:12px;font-family:Arial,sans-serif;margin-top:8px;">
  <tr><td colspan="2" style="background:#E8F5E9;padding:7px 10px;border:1px solid #CCCCCC;font-weight:bold;">
    📊 Excel attached — 3 sheets
  </td></tr>
  <tr><td style="padding:6px 10px;border:1px solid #DDDDDD;background:{LGRAY};width:36%;"><strong>Sheet 1 — TOP MOVERS</strong></td>
      <td style="padding:6px 10px;border:1px solid #DDDDDD;background:{LGRAY};">All qualifying candidates ranked by scanner score</td></tr>
  <tr><td style="padding:6px 10px;border:1px solid #DDDDDD;"><strong>Sheet 2 — TOP OPPORTUNITIES</strong></td>
      <td style="padding:6px 10px;border:1px solid #DDDDDD;">Top 10 ranked by entry quality (A+ first)</td></tr>
  <tr><td style="padding:6px 10px;border:1px solid #DDDDDD;background:{LGRAY};"><strong>Sheet 3 — DETAIL CARDS</strong></td>
      <td style="padding:6px 10px;border:1px solid #DDDDDD;background:{LGRAY};">Full TA: entry zone / stop / T1 / T2 / exit / sizing</td></tr>
</table>
<p style="font-family:Arial,sans-serif;font-size:11px;color:#888888;margin-top:10px;">
  Tier: 🔥 A+ highest conviction &nbsp;|&nbsp; 🔵 A strong &nbsp;|&nbsp; 🟡 B standard<br>
  Direction: 📈 LONG &nbsp;|&nbsp; 📉 SHORT<br>
  Weekly: ▲ W↑ price above 20-week MA (confirmed uptrend) &nbsp;|&nbsp; ▼ W↓ price below 20-week MA (downtrend / caution for longs)<br>
  📰 = news published in last 24 h<br><br>
  ⚠️ <em>Research only. Not financial advice.</em>
</p>"""

    return f"""<!DOCTYPE html>
<html lang="en">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0"></head>
<body style="margin:0;padding:0;background:#EFEFEF;font-family:Arial,sans-serif;">
  <table width="100%" cellpadding="0" cellspacing="0" style="background:{NAVY};padding:18px 32px;">
    <tr><td>
      <div style="color:#FFFFFF;font-size:20px;font-weight:bold;">📡 Intraday Scanner</div>
      <div style="color:#A8BFDA;font-size:12px;margin-top:3px;">{run_time}</div>
    </td></tr>
  </table>
  <table width="100%" cellpadding="0" cellspacing="0" style="max-width:1050px;margin:20px auto;">
    <tr><td style="background:#FFFFFF;padding:24px 28px;border-radius:6px;box-shadow:0 1px 4px rgba(0,0,0,0.10);">
      <h3 style="font-family:Arial,sans-serif;font-size:13px;font-weight:bold;color:{NAVY};
                 margin:0 0 12px 0;border-bottom:1px solid {DGRAY};padding-bottom:4px;">
        Top Opportunities — Ranked by Entry Quality &amp; Conviction Tier
      </h3>
      {table}
      <hr style="border:none;border-top:1px solid {DGRAY};margin:22px 0;">
      {legend}
    </td></tr>
  </table>
  <table width="100%" cellpadding="0" cellspacing="0" style="max-width:1050px;margin:0 auto 28px;">
    <tr><td style="text-align:center;font-size:10px;color:#AAAAAA;padding:8px;">
      JLG Hunt Bot &nbsp;·&nbsp; intraday-scanner &nbsp;·&nbsp; {run_time}
    </td></tr>
  </table>
</body>
</html>"""


def send_email_report(top10: list, filepath: str, run_time: str) -> bool:
    """
    Send HTML email with the opportunity table + Excel attachment.
    Required env vars: EMAIL_TO, EMAIL_FROM, EMAIL_PASSWORD
    """
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.base      import MIMEBase
    from email.mime.text      import MIMEText
    from email                import encoders

    email_to   = os.getenv("EMAIL_TO",       "")
    email_from = os.getenv("EMAIL_FROM",     "")
    email_pass = os.getenv("EMAIL_PASSWORD", "")
    smtp_host  = os.getenv("SMTP_HOST",      "smtp.gmail.com")
    smtp_port  = int(os.getenv("SMTP_PORT",  "587"))

    if not email_to or not email_from or not email_pass:
        print("WARN: EMAIL_TO / EMAIL_FROM / EMAIL_PASSWORD not set — skipping email")
        return False

    html_body = build_html_email(top10, run_time)

    msg            = MIMEMultipart("mixed")
    msg["From"]    = email_from
    msg["To"]      = email_to
    msg["Subject"] = f"Intraday Scanner — {run_time}"

    alt = MIMEMultipart("alternative")
    alt.attach(MIMEText(html_body, "html", "utf-8"))
    msg.attach(alt)

    with open(filepath, "rb") as fh:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(fh.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition",
                    f'attachment; filename="{os.path.basename(filepath)}"')
    msg.attach(part)

    try:
        srv = smtplib.SMTP(smtp_host, smtp_port)
        srv.ehlo()
        srv.starttls()
        srv.login(email_from, email_pass)
        srv.sendmail(email_from, email_to, msg.as_string())
        srv.quit()
        print(f"HTML email sent → {email_to}")
        return True
    except Exception as e:
        print(f"ERROR sending email: {e}")
        return False


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="TA Entry Runner")
    group  = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--tickers",   nargs="+", help="Space-separated tickers")
    group.add_argument("--from-file", help="Path to scan_results.json")
    args = parser.parse_args()

    if args.tickers:
        tickers      = args.tickers
        scanner_data = {}
    else:
        with open(args.from_file) as f:
            scan_data = json.load(f)
        tickers      = [d["ticker"] for d in scan_data]
        scanner_data = {d["ticker"]: d for d in scan_data}

    now_str = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    print(f"\n{'='*60}")
    print(f"TA ENTRY RUNNER -- {now_str}")
    print(f"Processing {len(tickers)} ticker(s): {', '.join(tickers)}")
    print(f"{'='*60}\n")

    results    = []
    actionable = []

    print(f"  Running parallel TA analysis (workers=10)...")
    with ThreadPoolExecutor(max_workers=10) as ex:
        fut_map = {ex.submit(run_ta_entry, t, scanner_data.get(t, {})): t for t in tickers}
        for fut in as_completed(fut_map):
            r = fut.result()
            results.append(r)
            ticker = r["ticker"]
            if r["status"] == "ok":
                print(f"  -> {ticker}: OK | {r['verdict'][:50]}")
                actionable.append(r)
            elif r["status"] == "pass":
                print(f"  -> {ticker}: PASS -- {r['pass_reason']}")
            else:
                print(f"  -> {ticker}: ERROR -- {r['error']}")

    # Rank by return probability, take top N
    top10 = rank_by_return_probability(actionable, top_n=TOP_N)

    print(f"\n  Ranked top {len(top10)} from {len(actionable)} actionable:")
    for i, r in enumerate(top10, 1):
        print(f"    {i}. {r['ticker']} | {r.get('trend_primary','-')} | {r.get('verdict','-')[:50]}")

    # Improvement 11: correlation / overlap guard
    compute_correlation_flags(top10)
    corr_pairs = [(r["ticker"], r["corr_flag"]) for r in top10 if r.get("corr_flag")]
    if corr_pairs:
        print("  Correlation flags (|corr| ≥ 0.75):")
        for t, others in corr_pairs:
            print(f"    {t} ↔ {', '.join(others)}")

    # Write orders.json — top 10 only, ranked
    orders = [
        {
            "rank":             i + 1,
            "ticker":           r["ticker"],
            "name":             r.get("name", r["ticker"]),
            "currency":         r["currency"],
            "timestamp_utc":    datetime.utcnow().isoformat(),
            "current_price":    r["current_price"],
            "entry_limit_low":  r["entry_low"],
            "entry_limit_high": r["entry_high"],
            "stop_loss":        r["stop_loss"],
            "stop_pct":         r["stop_pct"],
            "target_1":         r["target_1"],
            "rr_t1":            r["rr_t1"],
            "target_2":         r["target_2"],
            "rr_t2":            r["rr_t2"],
            "recommended_exit": r["recommended_exit"],
            "rr_exit":          r["rr_exit"],
            "verdict":          r["verdict"],
            "catalyst":         r.get("catalyst_note"),
            "rsi":              r["rsi"],
            "ma50":             r["ma50"],
            "ma200":            r["ma200"],
            "week52_low":       r["week52_low"],
            "week52_high":      r["week52_high"],
            # Enriched fields for ta_narrative.py
            "bias":             r.get("bias", "LONG"),
            "catalyst_tier":    r.get("catalyst_tier", "B"),
            "trend_primary":    r.get("trend_primary", ""),
            "trend_secondary":  r.get("trend_secondary", ""),
            "momentum_st":      r.get("momentum_st", ""),
            "adx":              r.get("adx"),
            "day_return":       r.get("day_return"),
            "rvol":             r.get("rvol"),
            "rs_vs_bench":      r.get("rs_vs_bench"),
            # Improvements 1, 3, 4, 5, 6, 7, 8, 9, 10, 11
            "has_news_today":        r.get("has_news_today", False),
            "news_headline":         r.get("news_headline", ""),
            "weekly_trend_up":       r.get("weekly_trend_up"),
            "rs_vs_sector":          r.get("rs_vs_sector"),
            "short_interest_pct":    r.get("short_interest_pct"),
            "earnings_surprise_avg": r.get("earnings_surprise_avg"),
            "futures_roll_warning":  r.get("futures_roll_warning", False),
            "days_to_roll":          r.get("days_to_roll"),
            "roll_date":             r.get("roll_date"),
            "repeat_days":           r.get("repeat_days", 0),
            "score_trend":           r.get("score_trend", "→"),
            "sizing_capped_by":      r.get("sizing_capped_by"),
            "corr_flag":             r.get("corr_flag", []),
        }
        for i, r in enumerate(top10)
    ]
    with open("orders.json", "w") as f:
        json.dump(orders, f, indent=2)
    print(f"\norders.json written -- {len(orders)} ranked setup(s)")

    # Write ta_report.md — full audit trail of all tickers
    md = [
        f"# TA Entry Report -- {now_str}",
        f"",
        f"**{len(actionable)} actionable** from {len(tickers)} scanned | "
        f"**Top {len(top10)} ranked** by return probability",
        f"",
        f"| Ticker | Price | Entry Zone | Stop | T1 (R/R) | Exit (R/R) | Verdict |",
        f"|--------|-------|------------|------|----------|-----------|---------|",
    ]
    for r in results:
        md.append(format_markdown_row(r))

    with open("ta_report.md", "w") as f:
        f.write("\n".join(md))
    print("ta_report.md written")

    # Send Telegram — summary list first, then individual cards
    if top10:
        send_telegram(format_best_opportunities_summary(top10, len(actionable), len(tickers)))
        for i, r in enumerate(top10, 1):
            send_telegram(format_telegram_card(r, rank=i))
    else:
        send_telegram(
            f"🎯 *BEST OPPORTUNITIES — {now_str}*\n"
            f"0 actionable setups from {len(tickers)} scanned."
        )

    # ── Excel report + HTML email ─────────────────────────────────────────────
    xl_path        = f"scan_report_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    all_candidates = list(scanner_data.values()) if scanner_data else []
    try:
        generate_excel_report(all_candidates, top10, filepath=xl_path)
        print(f"Excel report written: {xl_path}")
        send_email_report(top10, xl_path, now_str)
    except ImportError:
        print("WARN: openpyxl not installed — skipping Excel report")
    except Exception as e:
        print(f"ERROR generating Excel/email report: {e}")

    print(f"\n{'='*60}")
    print(f"Done. {len(actionable)}/{len(tickers)} actionable. Top {len(top10)} sent to Telegram.")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
